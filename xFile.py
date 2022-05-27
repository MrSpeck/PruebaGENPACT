#Abrir el archivo (xFile)
#           ->resives la ruta del archivo de fileWatcher
#           ->copias cada sheet a el master file
#           ->mandas el archivo a la carpeta Processed

import openpyxl
import shutil

class XFile:

    def __init__(self,slaveFilePath, masterFilePath):
        self.slaveFilePath = slaveFilePath
        self.masterFilePath = masterFilePath

    def copySheets(self):
        masterWB = openpyxl.load_workbook(self.masterFilePath)
        slaveWB = openpyxl.load_workbook(self.slaveFilePath)
        masterSheetNumber = len(masterWB.sheetnames)
        slaveNames = slaveWB.sheetnames

        for i in slaveNames:
            maxRow = slaveWB[i].max_row
            maxCol = slaveWB[i].max_column
            masterSheetName = 'Hoja' + str(masterSheetNumber + 1)
            masterSheetNumber+=1
            masterWB.create_sheet(masterSheetName)

            for r in range (1, maxRow + 1):
                for c in range (1, maxCol + 1):
                    masterWB[masterSheetName].cell(row = r, column = c).value = slaveWB[i].cell(row = r, column = c).value

        masterWB.save(self.masterFilePath)

    def sendFile(self, processedPath):
        shutil.move(self.slaveFilePath, processedPath, copy_function = shutil.copy)
 
